"""
Buduje warframe_demand_tracker.xlsx.

Market Demand Index NIE jest wpisywany jako gotowa liczba z Pythona - jest formułą Excela,
która liczy logarytmiczną normalizację na żywo z kolumn J/K/F i wag z P2:Q4. Dzięki temu:
  - jeśli dopiszesz/usuniesz wiersz i przeliczysz zakres, indeks się zaktualizuje
  - możesz zmienić wagi bezpośrednio w Excelu (żółte komórki) bez odpalania Pythona

Kolory: niebieski tekst = dane wejściowe (hardcoded), czarny = formuły,
żółte tło = komórki do edycji (wag).
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from config import TRADABLE_ONLY
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)         # niebieski = hardcoded input
FORMULA_FONT = Font(name=FONT_NAME, color="000000", size=10)       # czarny = formuła
NOTE_FONT = Font(name=FONT_NAME, color="000000", size=9, italic=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
THIN_BORDER = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

COLUMNS = [
    ("Rank", 6),
    ("Item", 22),
    ("Category", 14),
    ("Lowest Sell (plat)", 15),
    ("Avg Top5 Sell (plat)", 16),
    ("Buy Orders Count", 14),
    ("Sell Orders Count", 14),
    ("Volume 48h", 12),
    ("Volume 90d avg/day", 16),
    ("Sales Count", 14),
    ("Avg Sell Price", 14),
    ("Crafting Uses Count", 16),
    ("Market Demand Index", 18),
    ("Grind Efficiency (plat/min)", 20),
    ("Used In (Crafting)", 40),
    ("Overframe Usage", 20),
    ("Acquisition Note", 55),
    ("Last Updated", 18),
]


def build_workbook(rows: list[dict], output_path: Path, report_name: str | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    report_title = report_name or "Demand Tracker"
    ws.title = report_title[:31]
    wb.properties.title = report_title
    wb.properties.subject = "Warframe demand tracker"

    # --- Header ---
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    n = len(rows)
    first_row, last_row = 2, n + 1

    # --- Data rows (hardcoded inputs, blue) ---
    now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, row in enumerate(rows):
        r = first_row + i
        values = [
            row.get("rank"),
            row.get("item_name"),
            row.get("category"),
            row.get("lowest_sell_price"),
            row.get("avg_sell_price_top5"),
            row.get("buy_orders_count", 0),
            row.get("sell_orders_count", 0),
            row.get("volume_48h", 0),
            row.get("volume_90d_avg", 0),
            row.get("sales_count", 0),
            row.get("avg_sell_price", 0),
            row.get("crafting_uses_count", 0),
            None,  # Market Demand Index - formuła, patrz niżej
            row.get("estimated_grind_yield"),
            row.get("crafting_uses_text", ""),
            row.get("overframe_usage_text", "N/A (Overframe wyłączony w config.py)"),
            row.get("acquisition_note", ""),
            now_str,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = INPUT_FONT if col_idx != 13 else FORMULA_FONT
            cell.border = THIN_BORDER
            if col_idx in (13, 15):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx in (4, 5, 11):
                cell.number_format = '#,##0"p"'
            if col_idx == 14:
                cell.number_format = '#,##0.00'

    # --- Market Demand Index formula (kolumna M = 13), logarytmiczna normalizacja na żywo ---
    def norm(col_letter: str, r: int) -> str:
        rng = f"${col_letter}${first_row}:${col_letter}${last_row}"
        return (
            f"IF(MAX({rng})<=0,0,LOG10({col_letter}{r}+1)/LOG10(MAX({rng})+1)*100)"
        )

    for i in range(n):
        r = first_row + i
        formula = (
            f"={norm('J', r)}*$Q$2"
            f"+{norm('K', r)}*$Q$3"
            f"+{norm('F', r)}*$Q$4"
        )
        cell = ws.cell(row=r, column=13, value=formula)
        cell.font = FORMULA_FONT
        cell.number_format = "0.0"
        cell.border = THIN_BORDER

    # --- Panel wag (P/Q), żółte = edytowalne ---
    ws["P1"] = "Wagi Market Demand Index"
    ws["P1"].font = Font(name=FONT_NAME, bold=True, size=10)
    weight_labels = [
        ("Sales Velocity", 0.50),
        ("Avg Sell Price", 0.30),
        ("Buy Order Demand", 0.20),
    ]
    for i, (label, default) in enumerate(weight_labels):
        r = 2 + i
        ws.cell(row=r, column=16, value=label).font = FORMULA_FONT
        wcell = ws.cell(row=r, column=17, value=default)
        wcell.font = INPUT_FONT
        wcell.fill = YELLOW_FILL
        wcell.number_format = "0.00"
    ws["P6"] = "Suma wag (powinno = 1.00)"
    ws["P6"].font = NOTE_FONT
    ws["Q6"] = "=SUM(Q2:Q4)"
    ws["Q6"].font = FORMULA_FONT
    ws.column_dimensions["P"].width = 20
    ws.column_dimensions["Q"].width = 10

    # --- Conditional formatting: Market Demand Index jako color scale ---
    if n >= 1:
        ws.conditional_formatting.add(
            f"M{first_row}:M{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )

    # --- Legenda ---
    legend_row = last_row + 3
    ws.cell(row=legend_row, column=1, value="Legenda:").font = Font(name=FONT_NAME, bold=True, size=9)
    legend_lines = [
        "Niebieski tekst = dane wejściowe z warframe.market / WFCD warframe-items (nadpisywane przy każdym uruchomieniu main.py)",
        "Czarny tekst = formuła Excela (Market Demand Index liczy się na żywo z kolumn J/K/F i wag w Q2:Q4)",
        "Żółte komórki (Q2:Q4) = wagi Market Demand Index - możesz je zmienić ręcznie w Excelu, indeks przeliczy się automatycznie",
        "Overframe Usage = eksperymentalny sygnał z overframe.gg, wyłączony domyślnie (patrz README.md)",
        "Grind Efficiency (plat/min) to przybliżony wskaźnik oparty na średniej cenie oraz czasie grindowania z WFCD; traktuj jako orientacyjny.",
    ]
    if TRADABLE_ONLY:
        legend_lines.append("Arkusz zawiera tylko przedmioty tradowalne z warframe.market.")
    else:
        legend_lines.append("Arkusz może zawierać również przedmioty niehandlowalne.")
    for i, line in enumerate(legend_lines):
        ws.cell(row=legend_row + 1 + i, column=1, value=line).font = NOTE_FONT

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
